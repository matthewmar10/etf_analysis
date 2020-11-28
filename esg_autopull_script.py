import openpyxl
from openpyxl import Workbook
from selenium import webdriver
import time

browser = webdriver.Firefox()

# pulls the data for the designated ETF
def pulldata(ticker):
	ticker = ticker
	evaluation_metrics = []
	
	url = "https://finance.yahoo.com/quote/"+ticker+"/holdings?p="+ticker
	
	browser.get(url)
	time.sleep(5)
	
	pe_ratio = browser.find_element_by_css_selector("div.W\(48\%\):nth-child(2) > div:nth-child(1) > div:nth-child(2) > div:nth-child(2) > span:nth-child(2)").text
	
	pb_ratio = browser.find_element_by_css_selector("div.W\(48\%\):nth-child(2) > div:nth-child(1) > div:nth-child(2) > div:nth-child(3) > span:nth-child(2)").text
	
	stock_percentage = browser.find_element_by_css_selector("div.W\(48\%\):nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(1) > span:nth-child(2)").text
	
	url = "https://finance.yahoo.com/quote/"+ticker+"?p="+ticker
	browser.get(url)
	time.sleep(6)
	
	expense_ratio = browser.find_element_by_css_selector("table.M\(0\):nth-child(1) > tbody:nth-child(1) > tr:nth-child(7) > td:nth-child(2) > span:nth-child(1)").text

	div_yield = browser.find_element_by_css_selector("table.M\(0\):nth-child(1) > tbody:nth-child(1) > tr:nth-child(4) > td:nth-child(2) > span:nth-child(1)").text

	beta = browser.find_element_by_css_selector("table.M\(0\) > tbody:nth-child(1) > tr:nth-child(6) > td:nth-child(2) > span:nth-child(1)").text
	
	#aum = 
	#time.sleep(5)
	
	print(ticker)
	print(pe_ratio)
	print(pb_ratio)
	
	if stock_percentage != "N/A":
		stock_percentage = stock_percentage[:-1]
		stock_percentage = float(stock_percentage)/100
	print(stock_percentage)
	
	if expense_ratio != "N/A":
		expense_ratio = expense_ratio[:-1]
		expense_ratio = float(expense_ratio)/100
	print(expense_ratio)

	if div_yield != "N/A":
		div_yield = div_yield[:-1]
		div_yield = float(div_yield)/100
	print(div_yield)

	print(beta)

	evaluation_metrics.append(pe_ratio)
	evaluation_metrics.append(pb_ratio)
	evaluation_metrics.append(stock_percentage)
	evaluation_metrics.append(expense_ratio)
	evaluation_metrics.append(div_yield)
	evaluation_metrics.append(beta)

	return evaluation_metrics



def pulldata_differently(ticker):
	ticker = ticker
	evaluation_metrics = []
	
	url = "https://finance.yahoo.com/quote/"+ticker+"/holdings?p="+ticker
	
	browser.get(url)
	time.sleep(5)
	
	pe_ratio = browser.find_element_by_css_selector("div.W\(48\%\):nth-child(2) > div:nth-child(1) > div:nth-child(2) > div:nth-child(2) > span:nth-child(2)").text
	
	pb_ratio = browser.find_element_by_css_selector("div.W\(48\%\):nth-child(2) > div:nth-child(1) > div:nth-child(2) > div:nth-child(3) > span:nth-child(2)").text
	
	stock_percentage = browser.find_element_by_css_selector("div.W\(48\%\):nth-child(1) > div:nth-child(1) > div:nth-child(2) > div:nth-child(2) > span:nth-child(2)").text
	
	url = "https://finance.yahoo.com/quote/"+ticker+"?p="+ticker
	browser.get(url)
	time.sleep(6)
	
	expense_ratio = browser.find_element_by_css_selector("div.W\(1\/2\):nth-child(1) > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child(3) > td:nth-child(2) > span:nth-child(1)").text

	div_yield = browser.find_element_by_css_selector("table.M\(0\):nth-child(1) > tbody:nth-child(1) > tr:nth-child(3) > td:nth-child(2) > span:nth-child(1)").text

	beta = browser.find_element_by_css_selector("table.M\(0\):nth-child(1) > tbody:nth-child(1) > tr:nth-child(2) > td:nth-child(2) > span:nth-child(1)").text
	
	#aum = 
	#time.sleep(5)
	
	print(ticker)
	print(pe_ratio)
	print(pb_ratio)
	
	if stock_percentage != "N/A":
		stock_percentage = stock_percentage[:-1]
		stock_percentage = float(stock_percentage)/100
	print(stock_percentage)
	
	if expense_ratio != "N/A":
		expense_ratio = expense_ratio[:-1]
		expense_ratio = float(expense_ratio)/100
	print(expense_ratio)

	if div_yield != "N/A":
		div_yield = div_yield[:-1]
		div_yield = float(div_yield)/100
	print(div_yield)

	print(beta)

	evaluation_metrics.append(pe_ratio)
	evaluation_metrics.append(pb_ratio)
	evaluation_metrics.append(stock_percentage)
	evaluation_metrics.append(expense_ratio)
	evaluation_metrics.append(div_yield)
	evaluation_metrics.append(beta)

	return evaluation_metrics
	


read_book = openpyxl.load_workbook("ESG_funds.xlsx")
read_sheet = read_book.active

dict = {}

# adds every ETF from the spreadsheet to a dictionary, with the tickers as the keys and evaluation metrics as the values
for i in range(4, 20, 1):
	read_cell = read_sheet.cell(row=i, column=1)
	
	ticker = read_cell.value

	evaluation_metrics = pulldata(ticker)
	
	dict[ticker] = evaluation_metrics


# adds the next ETFs, which have different Yahoo Finance layouts
for i in range(21, 24, 1):
	read_cell = read_sheet.cell(row=i, column=1)
	
	ticker = read_cell.value

	evaluation_metrics = pulldata_differently(ticker)
	
	dict[ticker] = evaluation_metrics


print(dict)



# add values to new Excel
write_book = Workbook()
write_sheet = write_book.active
write_row = 4

for i in dict:
	write_sheet.cell(row=write_row, column=1).value = i
	
	num = 2
	for x in dict[i]:
		write_sheet.cell(row=write_row, column=num).value = x
		num += 1
	
	write_row += 1

write_book.save("ESG_update.xlsx")

	
	


